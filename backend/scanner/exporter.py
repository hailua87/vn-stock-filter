"""
Export scanner results to Excel, JSON and HTML.
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import json
import pandas as pd


def to_excel(df: pd.DataFrame, output_path: str | Path, include_guide: bool = True):
    """Export results to a formatted Excel file."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Signals', index=False)
        ws = writer.sheets['Signals']

        # Header style
        header_fill = PatternFill('solid', start_color='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)),
                          df[col_name].astype(str).str.len().max() if len(df) > 0 else 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

        # Conditional formatting for rating
        if 'rating' in df.columns:
            col_idx = list(df.columns).index('rating') + 1
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if v == 'A+':
                    cell.fill = PatternFill('solid', start_color='00B050')
                    cell.font = Font(bold=True, color='FFFFFF')
                elif v == 'A':
                    cell.fill = PatternFill('solid', start_color='92D050')
                elif v == 'B':
                    cell.fill = PatternFill('solid', start_color='FFC000')
                else:
                    cell.fill = PatternFill('solid', start_color='D9D9D9')

        ws.freeze_panes = 'A2'

        if include_guide:
            _add_guide_sheet(writer)


def _add_guide_sheet(writer):
    """Add a documentation sheet explaining the criteria."""
    guide = pd.DataFrame({
        'Tiêu chí': [
            'c1_atr_squeeze', 'c2_bb_squeeze', 'c3_near_high20',
            'c4_stealth_accum', 'c5_vol_surge', 'c6_upper_close',
            'c7_ma_align', 'c8_rsi_zone', 'c9_pocket_pivot', 'c10_no_gap_down',
        ],
        'Tên': [
            'ATR siết', 'Bollinger Squeeze', 'Gần đỉnh 20 phiên',
            'Stealth Accumulation', 'Volume surge', 'Đóng cửa nửa trên',
            'MA10>MA20', 'RSI 50-65', 'Pocket Pivot', 'Không gap down',
        ],
        'Ý nghĩa': [
            'Biên độ giá siết lại < 85% TB 20 phiên',
            'BB width thuộc 25% thấp nhất 60 phiên',
            'Giá cách đỉnh 20 phiên < 3%, chưa break',
            'OBV tăng > 5% trong khi giá tăng < 3%',
            'Volume 5 phiên > 1.15 × MA20',
            'Đóng cửa ≥ 60% biên độ ngày, ít nhất 3/5 phiên',
            'MA10 > MA20 và MA20 hướng lên',
            'RSI(14) trong vùng 50-65',
            'Phiên tăng có volume vượt mọi phiên giảm 10 ngày',
            'Không có gap down > 4% trong 5 phiên',
        ],
        'Điểm': [1]*10,
    })
    guide.to_excel(writer, sheet_name='Hướng dẫn', index=False)


def to_json(df: pd.DataFrame, output_path: str | Path, metadata: dict | None = None):
    """Export to JSON for web consumption."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        'generated_at': datetime.now().isoformat(),
        'total': len(df),
        'metadata': metadata or {},
        'signals': df.to_dict('records'),
    }
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)


def to_html(df: pd.DataFrame, output_path: str | Path, title: str = "Pre-Breakout Signals"):
    """Standalone HTML report (no JS needed)."""
    html = f"""<!DOCTYPE html>
<html lang="vi"><head><meta charset="utf-8">
<title>{title}</title>
<style>
  body {{ font-family: -apple-system, sans-serif; padding: 20px; background: #0f1419; color: #e6e6e6; }}
  h1 {{ color: #4fc3f7; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #1f4e79; color: white; padding: 10px; text-align: left; position: sticky; top: 0; }}
  td {{ padding: 8px; border-bottom: 1px solid #2a2a2a; }}
  tr:hover {{ background: #1a2027; }}
  .a-plus {{ background: #00b050; color: white; padding: 2px 8px; border-radius: 4px; font-weight: bold; }}
  .a {{ background: #92d050; color: #002000; padding: 2px 8px; border-radius: 4px; }}
  .b {{ background: #ffc000; color: #4a2c00; padding: 2px 8px; border-radius: 4px; }}
</style></head><body>
<h1>📈 {title}</h1>
<p>Generated: {datetime.now():%Y-%m-%d %H:%M}</p>
{df.to_html(index=False, escape=False, classes='signals')}
</body></html>"""
    Path(output_path).write_text(html, encoding='utf-8')
